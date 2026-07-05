from __future__ import annotations

import csv
from pathlib import Path
from typing import Any


class FileLoader:
    """Load problem statements and data files into structured dictionaries."""

    SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt", ".csv", ".xlsx"}

    def load_problem(self, path: Path) -> dict[str, Any]:
        return self.load_file(path)

    def load_data(self, path: Path | None) -> list[dict[str, Any]]:
        if path is None:
            return []
        path = path.resolve()
        if path.is_file():
            return [self.load_file(path)]
        if not path.exists():
            return [
                {
                    "path": str(path),
                    "type": "missing",
                    "content": "",
                    "metadata": {},
                    "error": f"Data path does not exist: {path}",
                }
            ]
        loaded: list[dict[str, Any]] = []
        for file_path in sorted(path.rglob("*")):
            if file_path.is_file() and file_path.suffix.lower() in self.SUPPORTED_EXTENSIONS:
                loaded.append(self.load_file(file_path))
        return loaded

    def load_file(self, path: Path) -> dict[str, Any]:
        path = path.resolve()
        suffix = path.suffix.lower()
        base: dict[str, Any] = {
            "path": str(path),
            "name": path.name,
            "type": suffix.lstrip("."),
            "content": "",
            "tables": [],
            "metadata": {"size_bytes": path.stat().st_size if path.exists() else 0},
            "error": None,
        }
        if not path.exists():
            base["error"] = f"File does not exist: {path}"
            return base
        try:
            if suffix == ".txt":
                base["content"] = self._read_text(path)
            elif suffix == ".csv":
                table = self._read_csv(path)
                base["tables"] = [table]
                base["content"] = self._table_to_text(table)
            elif suffix == ".xlsx":
                base["tables"] = self._read_xlsx(path)
                base["content"] = "\n\n".join(self._table_to_text(t) for t in base["tables"])
            elif suffix == ".docx":
                base["content"] = self._read_docx(path)
            elif suffix == ".pdf":
                base["content"] = self._read_pdf(path)
            else:
                base["error"] = f"Unsupported extension: {suffix}"
        except Exception as exc:  # noqa: BLE001 - loaders should preserve recoverable errors.
            base["error"] = f"{type(exc).__name__}: {exc}"
        return base

    def _read_text(self, path: Path) -> str:
        for encoding in ("utf-8", "utf-8-sig", "gb18030"):
            try:
                return path.read_text(encoding=encoding)
            except UnicodeDecodeError:
                continue
        return path.read_text(errors="replace")

    def _read_csv(self, path: Path, max_rows: int = 200) -> dict[str, Any]:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(f, dialect=dialect)
            rows = []
            for idx, row in enumerate(reader):
                if idx >= max_rows:
                    break
                rows.append(dict(row))
        return {
            "name": path.name,
            "source_path": str(path),
            "columns": list(rows[0].keys()) if rows else [],
            "rows": rows,
            "row_count_preview": len(rows),
            "truncated": len(rows) >= max_rows,
        }

    def _read_xlsx(self, path: Path, max_rows: int = 200) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("openpyxl is required for .xlsx files. Install requirements.txt.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        tables: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            rows_iter = sheet.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if headers is None:
                tables.append(
                    {
                        "name": sheet.title,
                        "source_path": str(path),
                        "columns": [],
                        "rows": [],
                        "row_count_preview": 0,
                        "truncated": False,
                    }
                )
                continue
            columns = [str(v) if v is not None else f"column_{i + 1}" for i, v in enumerate(headers)]
            rows = []
            for idx, values in enumerate(rows_iter):
                if idx >= max_rows:
                    break
                rows.append({columns[i]: values[i] if i < len(values) else None for i in range(len(columns))})
            tables.append(
                {
                    "name": sheet.title,
                    "source_path": str(path),
                    "columns": columns,
                    "rows": rows,
                    "row_count_preview": len(rows),
                    "truncated": len(rows) >= max_rows,
                }
            )
        return tables

    def _read_docx(self, path: Path) -> str:
        try:
            from docx import Document
        except ImportError as exc:
            raise ImportError("python-docx is required for .docx files. Install requirements.txt.") from exc

        document = Document(path)
        parts = [p.text for p in document.paragraphs if p.text.strip()]
        return "\n".join(parts)

    def _read_pdf(self, path: Path) -> str:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise ImportError("pypdf is required for .pdf files. Install requirements.txt.") from exc

        reader = PdfReader(str(path))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        return "\n\n".join(pages)

    def _table_to_text(self, table: dict[str, Any]) -> str:
        lines = [f"Table: {table.get('name', '')}", "Columns: " + ", ".join(table.get("columns", []))]
        for row in table.get("rows", [])[:20]:
            lines.append(str(row))
        if table.get("truncated"):
            lines.append("... preview truncated ...")
        return "\n".join(lines)
