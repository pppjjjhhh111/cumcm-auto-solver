from __future__ import annotations

import csv
import math
import statistics
from collections import Counter
from pathlib import Path
from typing import Any

from src.utils.json_utils import write_json


class DataProfiler:
    """Profile csv/xlsx data files and generate lightweight SVG EDA figures."""

    SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
    TIME_HINTS = ("date", "time", "year", "month", "day", "日期", "时间", "年份", "月份")
    ID_HINTS = ("id", "编号", "代码", "code", "no")
    GEO_HINTS = ("lat", "lon", "lng", "province", "city", "region", "经度", "纬度", "省", "市", "地区")
    TARGET_HINTS = ("target", "label", "score", "value", "demand", "sales", "目标", "得分", "销量", "需求")

    def __init__(
        self,
        figures_dir: Path | None = None,
        logs_dir: Path | None = None,
        max_rows: int = 1000,
    ) -> None:
        self.figures_dir = figures_dir
        self.logs_dir = logs_dir
        self.max_rows = max_rows

    def run(self, data_path: Path | None) -> dict[str, Any]:
        profile = self.profile_path(data_path)
        if self.logs_dir is not None:
            write_json(self.logs_dir / "data_profile.json", profile)
        return profile

    def profile_path(self, data_path: Path | None) -> dict[str, Any]:
        if data_path is None:
            return self._empty_profile("No data path was provided.")

        data_path = data_path.resolve()
        if not data_path.exists():
            return self._empty_profile(f"Data path does not exist: {data_path}")

        files = self._discover_files(data_path)
        if not files:
            return self._empty_profile(f"No csv/xlsx files were found under: {data_path}")

        profile_files = []
        for file_path in files:
            try:
                profile_files.extend(self._profile_file(file_path))
            except Exception as exc:  # noqa: BLE001 - profiling should degrade without stopping workflow.
                profile_files.append(
                    {
                        "file_name": file_path.name,
                        "path": str(file_path),
                        "error": f"{type(exc).__name__}: {exc}",
                    }
                )

        result = {
            "agent": "DataProfiler",
            "status": "ok" if profile_files else "empty",
            "file_count": len(files),
            "table_count": len(profile_files),
            "files": profile_files,
            "warnings": [],
            "summary": self._summary(profile_files),
        }
        return result

    def _empty_profile(self, warning: str) -> dict[str, Any]:
        result = {
            "agent": "DataProfiler",
            "status": "skipped",
            "file_count": 0,
            "table_count": 0,
            "files": [],
            "warnings": [warning],
            "summary": {
                "numeric_columns": [],
                "categorical_columns": [],
                "time_columns": [],
                "recommended_preprocessing_steps": ["Proceed with text-only modeling assumptions."],
                "recommended_figures": [],
            },
        }
        if self.logs_dir is not None:
            write_json(self.logs_dir / "data_profile.json", result)
        return result

    def _discover_files(self, data_path: Path) -> list[Path]:
        if data_path.is_file() and data_path.suffix.lower() in self.SUPPORTED_EXTENSIONS:
            return [data_path]
        if data_path.is_dir():
            return sorted(
                path
                for path in data_path.rglob("*")
                if path.is_file() and path.suffix.lower() in self.SUPPORTED_EXTENSIONS
            )
        return []

    def _profile_file(self, file_path: Path) -> list[dict[str, Any]]:
        if file_path.suffix.lower() == ".csv":
            rows = self._read_csv(file_path)
            return [self._profile_table(file_path, file_path.name, rows)]
        tables = self._read_xlsx(file_path)
        return [
            self._profile_table(file_path, f"{file_path.name}::{sheet_name}", rows, sheet_name=sheet_name)
            for sheet_name, rows in tables
        ]

    def _read_csv(self, file_path: Path) -> list[dict[str, Any]]:
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(handle, dialect=dialect)
            return [dict(row) for _, row in zip(range(self.max_rows), reader)]

    def _read_xlsx(self, file_path: Path) -> list[tuple[str, list[dict[str, Any]]]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("openpyxl is required for xlsx profiling.") from exc

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        tables = []
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            headers = next(iterator, None)
            if headers is None:
                tables.append((sheet.title, []))
                continue
            columns = [str(value) if value is not None else f"column_{idx + 1}" for idx, value in enumerate(headers)]
            rows = []
            for _, values in zip(range(self.max_rows), iterator):
                rows.append({columns[idx]: values[idx] if idx < len(values) else None for idx in range(len(columns))})
            tables.append((sheet.title, rows))
        return tables

    def _profile_table(
        self,
        file_path: Path,
        display_name: str,
        rows: list[dict[str, Any]],
        sheet_name: str | None = None,
    ) -> dict[str, Any]:
        columns = list(rows[0].keys()) if rows else []
        column_types = {column: self._infer_column_type([row.get(column) for row in rows]) for column in columns}
        missing_values = {
            column: sum(1 for row in rows if self._is_missing(row.get(column)))
            for column in columns
        }
        numeric_columns = [column for column, kind in column_types.items() if kind == "numeric"]
        categorical_columns = [column for column, kind in column_types.items() if kind == "categorical"]
        numeric_summary = {
            column: self._numeric_summary([self._to_float(row.get(column)) for row in rows])
            for column in numeric_columns
        }
        categorical_summary = {
            column: self._categorical_summary([row.get(column) for row in rows])
            for column in categorical_columns
        }
        correlation_matrix = self._correlation_matrix(rows, numeric_columns)
        profile = {
            "file_name": display_name,
            "path": str(file_path.resolve()),
            "sheet_name": sheet_name,
            "shape": {"rows": len(rows), "columns": len(columns)},
            "columns": columns,
            "column_types": column_types,
            "missing_values": missing_values,
            "duplicate_rows": self._duplicate_rows(rows),
            "numeric_summary": numeric_summary,
            "categorical_summary": categorical_summary,
            "correlation_matrix": correlation_matrix,
            "possible_target_columns": self._possible_target_columns(columns, numeric_columns),
            "time_columns": self._hint_columns(columns, self.TIME_HINTS),
            "id_columns": self._hint_columns(columns, self.ID_HINTS),
            "geo_columns": self._hint_columns(columns, self.GEO_HINTS),
            "recommended_preprocessing_steps": self._recommended_preprocessing_steps(
                missing_values,
                numeric_columns,
                categorical_columns,
            ),
            "recommended_figures": [],
        }
        profile["recommended_figures"] = self._generate_figures(profile, rows)
        return profile

    def _infer_column_type(self, values: list[Any]) -> str:
        non_missing = [value for value in values if not self._is_missing(value)]
        if not non_missing:
            return "empty"
        numeric = sum(1 for value in non_missing if self._to_float(value) is not None)
        if numeric / len(non_missing) >= 0.8:
            return "numeric"
        unique_ratio = len({str(value) for value in non_missing}) / max(len(non_missing), 1)
        return "categorical" if unique_ratio <= 0.7 or len(non_missing) < 30 else "text"

    def _numeric_summary(self, values: list[float | None]) -> dict[str, Any]:
        clean = [value for value in values if value is not None and math.isfinite(value)]
        if not clean:
            return {"count": 0}
        sorted_values = sorted(clean)
        return {
            "count": len(clean),
            "mean": statistics.fmean(clean),
            "std": statistics.stdev(clean) if len(clean) > 1 else 0.0,
            "min": min(clean),
            "q1": self._quantile(sorted_values, 0.25),
            "median": self._quantile(sorted_values, 0.5),
            "q3": self._quantile(sorted_values, 0.75),
            "max": max(clean),
            "outlier_count": self._outlier_count(sorted_values),
        }

    def _categorical_summary(self, values: list[Any]) -> dict[str, Any]:
        clean = [str(value) for value in values if not self._is_missing(value)]
        counts = Counter(clean)
        return {
            "count": len(clean),
            "unique": len(counts),
            "top_values": [{"value": key, "count": count} for key, count in counts.most_common(10)],
        }

    def _correlation_matrix(self, rows: list[dict[str, Any]], columns: list[str]) -> dict[str, dict[str, float | None]]:
        matrix: dict[str, dict[str, float | None]] = {}
        for left in columns:
            matrix[left] = {}
            for right in columns:
                if left == right:
                    matrix[left][right] = 1.0
                else:
                    matrix[left][right] = self._pearson(rows, left, right)
        return matrix

    def _pearson(self, rows: list[dict[str, Any]], left: str, right: str) -> float | None:
        pairs = []
        for row in rows:
            x = self._to_float(row.get(left))
            y = self._to_float(row.get(right))
            if x is not None and y is not None and math.isfinite(x) and math.isfinite(y):
                pairs.append((x, y))
        if len(pairs) < 3:
            return None
        xs = [x for x, _ in pairs]
        ys = [y for _, y in pairs]
        mean_x = statistics.fmean(xs)
        mean_y = statistics.fmean(ys)
        numerator = sum((x - mean_x) * (y - mean_y) for x, y in pairs)
        denom_x = math.sqrt(sum((x - mean_x) ** 2 for x in xs))
        denom_y = math.sqrt(sum((y - mean_y) ** 2 for y in ys))
        if denom_x == 0 or denom_y == 0:
            return None
        return numerator / (denom_x * denom_y)

    def _duplicate_rows(self, rows: list[dict[str, Any]]) -> int:
        seen = set()
        duplicates = 0
        for row in rows:
            key = tuple(sorted((str(k), str(v)) for k, v in row.items()))
            if key in seen:
                duplicates += 1
            else:
                seen.add(key)
        return duplicates

    def _recommended_preprocessing_steps(
        self,
        missing_values: dict[str, int],
        numeric_columns: list[str],
        categorical_columns: list[str],
    ) -> list[str]:
        steps = []
        if any(count > 0 for count in missing_values.values()):
            steps.append("Handle missing values with imputation, deletion, or explicit missing indicators.")
        if numeric_columns:
            steps.append("Standardize or normalize numeric columns when models are scale-sensitive.")
            steps.append("Check numeric outliers with IQR or z-score diagnostics.")
        if categorical_columns:
            steps.append("Encode categorical columns for machine learning models and keep labels for reporting.")
        if not steps:
            steps.append("No major preprocessing issue was detected in the previewed rows.")
        return steps

    def _generate_figures(self, profile: dict[str, Any], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        if self.figures_dir is None:
            return []
        figure_dir = self.figures_dir / "data_profile"
        figure_dir.mkdir(parents=True, exist_ok=True)
        safe_stem = self._safe_name(profile["file_name"])
        figures = []

        missing_path = figure_dir / f"{safe_stem}_missing.svg"
        self._write_bar_svg(missing_path, profile["missing_values"], "Missing values")
        figures.append(self._figure_record("missing_values", missing_path, "Missing value count by column."))

        numeric_summary = profile.get("numeric_summary", {})
        first_numeric = next(iter(numeric_summary.keys()), None)
        if first_numeric:
            values = [self._to_float(row.get(first_numeric)) for row in rows]
            hist_path = figure_dir / f"{safe_stem}_{self._safe_name(first_numeric)}_histogram.svg"
            self._write_histogram_svg(hist_path, [v for v in values if v is not None], f"Distribution of {first_numeric}")
            figures.append(self._figure_record("histogram", hist_path, f"Distribution of numeric column {first_numeric}."))

        categorical_summary = profile.get("categorical_summary", {})
        first_categorical = next(iter(categorical_summary.keys()), None)
        if first_categorical:
            counts = {
                item["value"]: item["count"]
                for item in categorical_summary[first_categorical].get("top_values", [])[:10]
            }
            cat_path = figure_dir / f"{safe_stem}_{self._safe_name(first_categorical)}_frequency.svg"
            self._write_bar_svg(cat_path, counts, f"Frequency of {first_categorical}")
            figures.append(self._figure_record("category_frequency", cat_path, f"Frequency of {first_categorical}."))

        if len(numeric_summary) >= 2:
            heatmap_path = figure_dir / f"{safe_stem}_correlation_heatmap.svg"
            self._write_heatmap_svg(heatmap_path, profile["correlation_matrix"], "Correlation heatmap")
            figures.append(self._figure_record("correlation_heatmap", heatmap_path, "Numeric correlation matrix."))
        return figures

    def _figure_record(self, figure_type: str, path: Path, caption: str) -> dict[str, Any]:
        return {"figure_type": figure_type, "path": str(path.resolve()), "caption": caption}

    def _write_bar_svg(self, path: Path, values: dict[str, Any], title: str) -> None:
        items = [(str(key), float(value or 0)) for key, value in list(values.items())[:12]]
        width = 900
        height = max(240, 90 + 28 * max(len(items), 1))
        max_value = max([value for _, value in items] or [1.0]) or 1.0
        rows = []
        for idx, (label, value) in enumerate(items):
            y = 60 + idx * 28
            bar_width = 680 * value / max_value if max_value else 0
            rows.append(
                f"<text x='20' y='{y + 15}' font-size='12' font-family='Arial'>{self._xml(label[:28])}</text>"
                f"<rect x='210' y='{y}' width='{bar_width:.1f}' height='18' fill='#2563eb'/>"
                f"<text x='{220 + bar_width:.1f}' y='{y + 15}' font-size='12' font-family='Arial'>{value:.4g}</text>"
            )
        svg = (
            f"<svg xmlns='http://www.w3.org/2000/svg' width='{width}' height='{height}'>"
            f"<rect width='100%' height='100%' fill='white'/>"
            f"<text x='20' y='32' font-size='20' font-family='Arial'>{self._xml(title)}</text>"
            + "".join(rows)
            + "</svg>"
        )
        path.write_text(svg, encoding="utf-8")

    def _write_histogram_svg(self, path: Path, values: list[float], title: str, bins: int = 10) -> None:
        if not values:
            path.write_text("<svg xmlns='http://www.w3.org/2000/svg'></svg>", encoding="utf-8")
            return
        min_value = min(values)
        max_value = max(values)
        span = max(max_value - min_value, 1e-9)
        counts = [0 for _ in range(bins)]
        for value in values:
            idx = min(int((value - min_value) / span * bins), bins - 1)
            counts[idx] += 1
        labels = {f"{min_value + span * i / bins:.2g}": count for i, count in enumerate(counts)}
        self._write_bar_svg(path, labels, title)

    def _write_heatmap_svg(self, path: Path, matrix: dict[str, dict[str, float | None]], title: str) -> None:
        columns = list(matrix.keys())[:12]
        cell = 34
        margin = 160
        width = margin + cell * len(columns) + 40
        height = margin + cell * len(columns) + 40
        parts = [
            f"<svg xmlns='http://www.w3.org/2000/svg' width='{width}' height='{height}'>",
            "<rect width='100%' height='100%' fill='white'/>",
            f"<text x='20' y='32' font-size='20' font-family='Arial'>{self._xml(title)}</text>",
        ]
        for i, row_name in enumerate(columns):
            parts.append(f"<text x='20' y='{margin + i * cell + 22}' font-size='11' font-family='Arial'>{self._xml(row_name[:18])}</text>")
            parts.append(f"<text x='{margin + i * cell}' y='145' font-size='10' font-family='Arial' transform='rotate(-45 {margin + i * cell},145)'>{self._xml(row_name[:18])}</text>")
            for j, col_name in enumerate(columns):
                corr = matrix.get(row_name, {}).get(col_name)
                value = 0.0 if corr is None else max(-1.0, min(1.0, float(corr)))
                color = self._corr_color(value)
                parts.append(
                    f"<rect x='{margin + j * cell}' y='{margin + i * cell}' width='{cell - 2}' height='{cell - 2}' fill='{color}'/>"
                )
        parts.append("</svg>")
        path.write_text("".join(parts), encoding="utf-8")

    def _corr_color(self, value: float) -> str:
        if value >= 0:
            blue = 180 - int(110 * value)
            return f"rgb({60},{130},{max(60, blue)})"
        red = 180 - int(110 * abs(value))
        return f"rgb({max(60, red)},{80},{80})"

    def _summary(self, profile_files: list[dict[str, Any]]) -> dict[str, Any]:
        numeric_columns = []
        categorical_columns = []
        time_columns = []
        steps = []
        figures = []
        for item in profile_files:
            for column, kind in item.get("column_types", {}).items():
                if kind == "numeric" and column not in numeric_columns:
                    numeric_columns.append(column)
                if kind == "categorical" and column not in categorical_columns:
                    categorical_columns.append(column)
            for column in item.get("time_columns", []):
                if column not in time_columns:
                    time_columns.append(column)
            for step in item.get("recommended_preprocessing_steps", []):
                if step not in steps:
                    steps.append(step)
            for figure in item.get("recommended_figures", []):
                figures.append(figure)
        return {
            "numeric_columns": numeric_columns,
            "categorical_columns": categorical_columns,
            "time_columns": time_columns,
            "recommended_preprocessing_steps": steps,
            "recommended_figures": figures,
        }

    def _possible_target_columns(self, columns: list[str], numeric_columns: list[str]) -> list[str]:
        hinted = [column for column in columns if any(hint in column.lower() for hint in self.TARGET_HINTS)]
        targets = hinted + [column for column in numeric_columns if column not in hinted]
        return targets[:5]

    def _hint_columns(self, columns: list[str], hints: tuple[str, ...]) -> list[str]:
        return [column for column in columns if any(hint in column.lower() for hint in hints)]

    def _is_missing(self, value: Any) -> bool:
        return value is None or str(value).strip() == ""

    def _to_float(self, value: Any) -> float | None:
        if self._is_missing(value):
            return None
        try:
            return float(str(value).replace(",", ""))
        except (TypeError, ValueError):
            return None

    def _quantile(self, sorted_values: list[float], q: float) -> float:
        if not sorted_values:
            return 0.0
        idx = (len(sorted_values) - 1) * q
        lower = math.floor(idx)
        upper = math.ceil(idx)
        if lower == upper:
            return sorted_values[int(idx)]
        return sorted_values[lower] * (upper - idx) + sorted_values[upper] * (idx - lower)

    def _outlier_count(self, sorted_values: list[float]) -> int:
        if len(sorted_values) < 4:
            return 0
        q1 = self._quantile(sorted_values, 0.25)
        q3 = self._quantile(sorted_values, 0.75)
        iqr = q3 - q1
        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr
        return sum(1 for value in sorted_values if value < lower or value > upper)

    def _safe_name(self, value: str) -> str:
        return "".join(ch if ch.isalnum() else "_" for ch in value)[:80].strip("_") or "data"

    def _xml(self, value: str) -> str:
        return value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
